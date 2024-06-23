import platform
import pandas as pd
import requests
from datetime import datetime, timedelta
from telegram import Update
from telegram.ext import Application, CommandHandler, CallbackContext, ConversationHandler, MessageHandler, filters
import pytz
from openpyxl import load_workbook
from opencage.geocoder import OpenCageGeocode
import os
import logging
import subprocess

# Configure logging
logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    level=logging.INFO
)
logger = logging.getLogger(__name__)

# Define conversation state
LOCATION = 0

# Function to get sunrise and sunset times
def get_sun_times(lat, lng, local_tz):
    today = datetime.now().strftime('%Y-%m-%d')
    tomorrow = (datetime.now() + timedelta(days=1)).strftime('%Y-%m-%d')

    url = f'https://api.sunrise-sunset.org/json?lat={lat}&lng={lng}&formatted=0&date='

    today_response = requests.get(url + today).json()
    tomorrow_response = requests.get(url + tomorrow).json()

    sunrise_today_utc = today_response['results']['sunrise']
    sunset_today_utc = today_response['results']['sunset']
    sunrise_tomorrow_utc = tomorrow_response['results']['sunrise']

    # Convert to local time
    ist = pytz.timezone(local_tz)
    sunrise_today = pd.to_datetime(sunrise_today_utc).tz_convert(ist)
    sunset_today = pd.to_datetime(sunset_today_utc).tz_convert(ist)
    sunrise_tomorrow = pd.to_datetime(sunrise_tomorrow_utc).tz_convert(ist)

    return sunrise_today, sunset_today, sunrise_tomorrow

# Function to update Excel file
def update_excel(file_path, sunrise_today, sunset_today, sunrise_tomorrow):
    try:
        wb = load_workbook(file_path)
        ws = wb.active
        
        ws['O4'] = sunrise_today.strftime('%H:%M:%S')
        ws['O5'] = sunset_today.strftime('%H:%M:%S')
        ws['O6'] = sunrise_tomorrow.strftime('%H:%M:%S')

        wb.save(file_path)
        wb.close()
        logger.info(f'Excel file {file_path} updated successfully.')
    except Exception as e:
        logger.error(f'Error updating Excel file: {e}')

# Function to convert Excel to image using JavaScript
def save_excel_range_as_image(excel_file_path, output_image_path, sheet_name):
    try:
        result = subprocess.run(
            ['node', 'excel_to_image.js', excel_file_path, sheet_name, output_image_path],
            capture_output=True, text=True
        )
        if result.returncode == 0:
            logger.info(f"Image successfully saved to {output_image_path}")
        else:
            logger.error(f"Error converting Excel to image: {result.stderr}")
    except Exception as e:
        logger.error(f"Error running JavaScript script: {e}")

# Command handler to start the conversation
async def send_table_start(update: Update, context: CallbackContext):
    await update.message.reply_text("Please enter your location (Vijayawada):")
    return LOCATION

# Function to handle location input and send the table with image
async def receive_location(update: Update, context: CallbackContext):
    location = update.message.text
    logger.info(f'Received location: {location}')

    # Use OpenCage Geocoder to get coordinates
    geocoder = OpenCageGeocode(context.bot_data['opencage_api_key'])
    result = geocoder.geocode(location)

    if result and len(result):
        latitude = result[0]['geometry']['lat']
        longitude = result[0]['geometry']['lng']
        logger.info(f'Coordinates for {location}: {latitude}, {longitude}')
        
        local_tz = 'Asia/Kolkata'  # Assuming Indian Standard Time (IST)
        
        # Get sun times
        sunrise_today, sunset_today, sunrise_tomorrow = get_sun_times(latitude, longitude, local_tz)

        # Update the Excel file
        file_path = context.bot_data['excel_file_path']
        update_excel(file_path, sunrise_today, sunset_today, sunrise_tomorrow)

        # Save Excel range as image
        save_image_path = context.bot_data['image_save_path']
        save_excel_range_as_image(file_path, save_image_path, "Sheet1")

        # Send the saved image
        await update.message.reply_photo(photo=open(save_image_path, 'rb'))

        # End the conversation
        return ConversationHandler.END
    else:
        logger.warning(f'Could not find coordinates for location: {location}')
        await update.message.reply_text("Sorry, I couldn't find coordinates for that location. Please try again.")
        return LOCATION

async def help_command_handler(update: Update, context: CallbackContext):
    await update.message.reply_text("This is the help message.")

async def main_handler(update: Update, context: CallbackContext):
    await update.message.reply_text("You sent a text message.")

def main():
    # Prompt user to enter tokens and paths
    opencage_api_key = '699522e909454a09b82d1c728fc79925'
    excel_file_path = r'/app/Completed_eng_Bharghava_Siddhanta_Panchangam.xlsx'
    image_save_path = r'/app/ExcelToImage.png'
    bot_token  ='7274941037:AAHIWiU5yvfIzo7eJWPu9S5CeJIid6ATEyM'

    # Create the Application instance
    application = Application.builder().token(bot_token).build()

    # Save tokens and paths in bot_data
    application.bot_data['opencage_api_key'] = opencage_api_key
    application.bot_data['excel_file_path'] = excel_file_path
    application.bot_data['image_save_path'] = image_save_path

    # Create the conversation handler
    conversation_handler = ConversationHandler(
        entry_points=[CommandHandler('panchangam', send_table_start)],
        states={
            LOCATION: [MessageHandler(filters.TEXT & ~filters.COMMAND, receive_location)],
        },
        fallbacks=[]
    )

    # Add conversation handler
    application.add_handler(conversation_handler)

    # Add other handlers
    application.add_handler(CommandHandler("help", help_command_handler))
    application.add_handler(MessageHandler(filters.TEXT, main_handler))

    # Run the bot
    application.run_polling()
    logger.info('Bot started successfully.')

if __name__ == '__main__':
    main()
